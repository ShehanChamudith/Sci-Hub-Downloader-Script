import os
import time
import random
import pandas as pd
import requests
from tqdm import tqdm
from PyPDF2 import PdfReader

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ─────────────────────────────────────────────
# 📄 Load DOIs and Titles
print("📄 Loading DOIs and Titles from CSV...")
CSV_FILE = "Papers for the SLR - Shehan.csv"
df = pd.read_csv(CSV_FILE)
df = df.dropna(subset=["DOI", "Title"])
entries = list(df[["DOI", "Title"]].itertuples(index=False, name=None))
print(f"✅ Loaded {len(entries)} entries.\n")

# 📂 Prepare download folder
BASE_DIR = os.path.dirname(os.path.abspath(CSV_FILE))
DOWNLOAD_DIR = os.path.join(BASE_DIR, "Script Downloads")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# ─────────────────────────────────────────────
# 🌐 Setup Chrome driver
print("🌐 Setting up Chrome browser...")
options = uc.ChromeOptions()
options.add_argument("--no-sandbox")
options.add_argument("--disable-blink-features=AutomationControlled")
driver = uc.Chrome(headless=False, options=options)
print("✅ Chrome setup complete.\n")

# ─────────────────────────────────────────────
# 🚀 Start scraping loop
downloaded = []
not_found = []

for i, (doi, title) in enumerate(entries, 1):
    short_title = " ".join(title.split()[:5])
    print(f"───────────────────────────────────────────────")
    print(f"🔍 [{i}/{len(entries)}] Searching for: {short_title}")

    try:
        time.sleep(random.uniform(0.5, 2.5))  # Random delay before navigation
        driver.get("https://sci-hub.se/")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "request")))

        input_box = driver.find_element(By.NAME, "request")
        input_box.clear()

        # Simulate human-like typing
        actions = ActionChains(driver)
        for char in doi:
            actions.send_keys(char)
            actions.pause(random.uniform(0.05, 0.15))  # Random pause per keystroke
        actions.perform()
        input_box.submit()

        # Random delay after submission
        time.sleep(random.uniform(0.5, 2))

        WebDriverWait(driver, 10).until(lambda d: "article not found" in d.title.lower() or "sci-hub" in d.title.lower())
        page_title = driver.title

        if "article not found" in page_title.lower():
            not_found.append(f"{title} | {doi}")
            print("❌ Paper not available.")
            continue

        print("📄 Paper found. Locating PDF embed...")

        try:
            embed = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "embed[type='application/pdf']"))
            )
            pdf_url = embed.get_attribute("src")

            if pdf_url.startswith("//"):
                pdf_url = "https:" + pdf_url
            elif pdf_url.startswith("/"):
                pdf_url = "https://sci-hub.se" + pdf_url

            print("📥 Downloading PDF...")

            headers = {
                "User-Agent": "Mozilla/5.0",
                "Referer": "https://sci-hub.se"
            }
            response = requests.get(pdf_url, headers=headers, stream=True, timeout=20)

            safe_title = "_".join(title.split()[:8]).replace("/", "_").replace("\\", "_")
            filename = safe_title + ".pdf"
            filepath = os.path.join(DOWNLOAD_DIR, filename)

            if response.status_code == 200 and "application/pdf" in response.headers.get("Content-Type", ""):
                with open(filepath, "wb") as f:
                    for chunk in tqdm(response.iter_content(1024), desc="💾 Saving", unit="KB", leave=False):
                        f.write(chunk)

                # 🧪 Validate PDF
                try:
                    reader = PdfReader(filepath)
                    _ = reader.pages[0]
                    downloaded.append(title)
                    print("✅ Download complete and PDF is valid.")
                except Exception as e:
                    not_found.append(f"{title} | {doi} | Corrupt PDF")
                    os.remove(filepath)
                    print(f"❌ Corrupt PDF: {e}")

            else:
                not_found.append(f"{title} | {doi} | Invalid content-type or status")
                debug_path = os.path.join(DOWNLOAD_DIR, f"debug_{safe_title}.html")
                with open(debug_path, "wb") as f:
                    f.write(response.content)
                print(f"❌ Invalid PDF (saved HTML to debug): {debug_path}")

        except TimeoutException:
            not_found.append(f"{title} | {doi}")
            print("❌ PDF embed not found.")

    except Exception as e:
        not_found.append(f"{title} | {doi}")
        print(f"❌ Error: {e}")

    # Optional random wait between entries
    time.sleep(random.uniform(1, 5))

# ─────────────────────────────────────────────
# 🧹 Cleanup
driver.quit()

if not_found:
    with open(os.path.join(DOWNLOAD_DIR, "not_found.txt"), "w", encoding="utf-8") as f:
        for entry in not_found:
            f.write(entry + "\n")

# 📊 Final Summary
print("\n📊 SUMMARY")
print(f"📥 Successfully downloaded: {len(downloaded)}")
print(f"❌ Not found or failed: {len(not_found)}")
print(f"📂 PDFs saved in: {DOWNLOAD_DIR}")
print("✅ Script finished.")

# 🗂 Update the original CSV with status and save as Excel
print("\n📁 Updating Excel with status results...")

# Reload to avoid previous modifications
df = pd.read_csv(CSV_FILE)

# Build DOI-based status map
downloaded_dois = set()
not_found_dois = set()

for doi, title in entries:
    if title in downloaded:
        downloaded_dois.add(doi)

for entry in not_found:
    parts = entry.split("|")
    if len(parts) >= 2:
        raw_doi = parts[1].strip()
        not_found_dois.add(raw_doi)

# Generate final status mapping using DOI
df["Status"] = df["DOI"].apply(
    lambda x: (
        "Found" if isinstance(x, str) and x.strip() in downloaded_dois else
        "Not Found" if isinstance(x, str) and x.strip() in not_found_dois else
        ""
    )
)


updated_excel = CSV_FILE.replace(".csv", "_status.xlsx")
df.to_excel(updated_excel, index=False)

# 🎨 Apply color formatting using openpyxl
wb = load_workbook(updated_excel)
ws = wb.active
status_col = None

for idx, cell in enumerate(ws[1], 1):
    if cell.value == "Status":
        status_col = idx
        break

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

for row in range(2, ws.max_row + 1):
    status = ws.cell(row=row, column=status_col).value
    if status == "Found":
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = green_fill
    elif status == "Not Found":
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = orange_fill

wb.save(updated_excel)
print(f"📊 Excel file saved with status and highlights: {updated_excel}")